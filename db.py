import sqlite3
import pandas as pd
from datetime import datetime
import os
import io as python_io

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "suivi_citernes.db")

def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def create_tables():
    conn = get_connection()
    cursor = conn.cursor()
    
    # Table Citernes
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS citernes (
        code TEXT PRIMARY KEY,
        type TEXT NOT NULL, -- 'CARBURANT' or 'EAU'
        comments TEXT DEFAULT '',
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
    """)
    
    # Table Etapes
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS etapes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        citerne_type TEXT NOT NULL,
        category TEXT NOT NULL,
        step_name TEXT NOT NULL,
        cadence_hours REAL DEFAULT 1.0,
        step_order INTEGER NOT NULL
    );
    """)
    
    # Table Suivi Progress
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS suivi_progress (
        citerne_code TEXT NOT NULL,
        step_id INTEGER NOT NULL,
        completion_pct REAL DEFAULT 0.0,
        comment TEXT DEFAULT '',
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        PRIMARY KEY (citerne_code, step_id),
        FOREIGN KEY (citerne_code) REFERENCES citernes (code) ON DELETE CASCADE,
        FOREIGN KEY (step_id) REFERENCES etapes (id) ON DELETE CASCADE
    );
    """)
    
    # Table Daily Snapshots for historical tracking
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS daily_snapshots (
        snapshot_date TEXT NOT NULL,
        citerne_code TEXT NOT NULL,
        step_id INTEGER NOT NULL,
        completion_pct REAL DEFAULT 0.0,
        PRIMARY KEY (snapshot_date, citerne_code, step_id)
    );
    """)
    
    conn.commit()
    conn.close()

def get_all_citernes(citerne_type=None):
    conn = get_connection()
    if citerne_type:
        df = pd.read_sql_query("SELECT * FROM citernes WHERE type = ? ORDER BY code", conn, params=(citerne_type,))
    else:
        df = pd.read_sql_query("SELECT * FROM citernes ORDER BY type, code", conn)
    conn.close()
    return df

def get_etapes(citerne_type):
    conn = get_connection()
    df = pd.read_sql_query("SELECT * FROM etapes WHERE citerne_type = ? ORDER BY step_order", conn, params=(citerne_type,))
    conn.close()
    return df

def get_citerne_details(citerne_code):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM citernes WHERE code = ?", (citerne_code,))
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None

def get_citerne_progress(citerne_code):
    conn = get_connection()
    query = """
    SELECT 
        e.id as step_id,
        e.category,
        e.step_name,
        e.cadence_hours,
        e.step_order,
        COALESCE(p.completion_pct, 0.0) as completion_pct,
        COALESCE(p.comment, '') as step_comment,
        p.updated_at
    FROM etapes e
    JOIN citernes c ON c.type = e.citerne_type
    LEFT JOIN suivi_progress p ON p.citerne_code = c.code AND p.step_id = e.id
    WHERE c.code = ?
    ORDER e.step_order
    """
    # Fix order by query: ORDER BY e.step_order
    query = query.replace("ORDER e.step_order", "ORDER BY e.step_order")
    df = pd.read_sql_query(query, conn, params=(citerne_code,))
    conn.close()
    return df

def update_step_progress(citerne_code, step_id, completion_pct, comment=None):
    conn = get_connection()
    cursor = conn.cursor()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    if comment is not None:
        cursor.execute("""
        INSERT INTO suivi_progress (citerne_code, step_id, completion_pct, comment, updated_at)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(citerne_code, step_id) DO UPDATE SET
            completion_pct = excluded.completion_pct,
            comment = excluded.comment,
            updated_at = excluded.updated_at
        """, (citerne_code, step_id, float(completion_pct), str(comment), now))
    else:
        cursor.execute("""
        INSERT INTO suivi_progress (citerne_code, step_id, completion_pct, comment, updated_at)
        VALUES (?, ?, ?, '', ?)
        ON CONFLICT(citerne_code, step_id) DO UPDATE SET
            completion_pct = excluded.completion_pct,
            updated_at = excluded.updated_at
        """, (citerne_code, step_id, float(completion_pct), now))
        
    cursor.execute("UPDATE citernes SET updated_at = ? WHERE code = ?", (now, citerne_code))
    conn.commit()
    conn.close()

def update_citerne_comment(citerne_code, comment):
    conn = get_connection()
    cursor = conn.cursor()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cursor.execute("UPDATE citernes SET comments = ?, updated_at = ? WHERE code = ?", (str(comment), now, citerne_code))
    conn.commit()
    conn.close()

def get_summary_dataframe(citerne_type=None):
    conn = get_connection()
    
    query_c = "SELECT code, type, comments, updated_at FROM citernes"
    if citerne_type:
        query_c += f" WHERE type = '{citerne_type}'"
    query_c += " ORDER BY type, code"
    
    citernes_df = pd.read_sql_query(query_c, conn)
    
    # Get total cadence per type
    cadence_df = pd.read_sql_query("SELECT citerne_type, SUM(cadence_hours) as total_cadence FROM etapes GROUP BY citerne_type", conn)
    cadence_dict = dict(zip(cadence_df['citerne_type'], cadence_df['total_cadence']))
    
    # Calculate progress for each citerne
    query_p = """
    SELECT 
        c.code,
        c.type,
        SUM((COALESCE(p.completion_pct, 0.0) / 100.0) * e.cadence_hours) as completed_hours,
        COUNT(CASE WHEN COALESCE(p.completion_pct, 0.0) >= 100 THEN 1 END) as steps_done,
        COUNT(e.id) as total_steps
    FROM citernes c
    JOIN etapes e ON e.citerne_type = c.type
    LEFT JOIN suivi_progress p ON p.citerne_code = c.code AND p.step_id = e.id
    """
    if citerne_type:
        query_p += f" WHERE c.type = '{citerne_type}'"
    query_p += " GROUP BY c.code, c.type"
    
    prog_df = pd.read_sql_query(query_p, conn)
    conn.close()
    
    merged = pd.merge(citernes_df, prog_df, on=['code', 'type'], how='left')
    merged['total_cadence'] = merged['type'].map(cadence_dict)
    merged['global_pct'] = (merged['completed_hours'] / merged['total_cadence'] * 100.0).round(1)
    merged['global_pct'] = merged['global_pct'].fillna(0.0)
    
    def get_status(row):
        pct = row['global_pct']
        if pct >= 99.9:
            return "Terminé"
        elif pct > 0:
            return "En cours"
        else:
            return "Non commencé"
            
    merged['statut'] = merged.apply(get_status, axis=1)
    return merged

def get_category_progress_df(citerne_type):
    conn = get_connection()
    query = """
    SELECT 
        e.category,
        AVG(COALESCE(p.completion_pct, 0.0)) as avg_completion,
        SUM((COALESCE(p.completion_pct, 0.0)/100.0) * e.cadence_hours) / SUM(e.cadence_hours) * 100.0 as weighted_completion
    FROM etapes e
    JOIN citernes c ON c.type = e.citerne_type
    LEFT JOIN suivi_progress p ON p.citerne_code = c.code AND p.step_id = e.id
    WHERE e.citerne_type = ?
    GROUP BY e.category
    ORDER BY MIN(e.step_order)
    """
    df = pd.read_sql_query(query, conn, params=(citerne_type,))
    conn.close()
    df['weighted_completion'] = df['weighted_completion'].round(1)
    df['avg_completion'] = df['avg_completion'].round(1)
    return df

def export_full_matrix(citerne_type):
    conn = get_connection()
    
    etapes = get_etapes(citerne_type)
    
    query = """
    SELECT 
        c.code as Citerne,
        e.step_order,
        COALESCE(p.completion_pct, 0.0) as completion
    FROM citernes c
    JOIN etapes e ON e.citerne_type = c.type
    LEFT JOIN suivi_progress p ON p.citerne_code = c.code AND p.step_id = e.id
    WHERE c.type = ?
    """
    p_df = pd.read_sql_query(query, conn, params=(citerne_type,))
    conn.close()
    
    # Pivot on unique step_order to prevent duplicates
    pivot = p_df.pivot(index='Citerne', columns='step_order', values='completion')
    
    # Rename columns to step_name
    step_names_dict = dict(zip(etapes['step_order'], etapes['step_name']))
    pivot = pivot.rename(columns=step_names_dict)
    
    return pivot

def write_cell_safely(ws, row, col, value, fill=None, font=None, alignment=None, border=None, number_format=None):
    cell = ws.cell(row=row, column=col)
    if type(cell).__name__ == 'MergedCell':
        for merged_range in ws.merged_cells.ranges:
            if merged_range.bounds[0] <= col <= merged_range.bounds[2] and \
               merged_range.bounds[1] <= row <= merged_range.bounds[3]:
                cell = ws.cell(row=merged_range.bounds[1], column=merged_range.bounds[0])
                break
                
    cell.value = value
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell

def generate_styled_excel_report(export_type="TOUS"):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    excel_template_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "SUIVI CITERNE 2107.xlsx")
    wb = openpyxl.load_workbook(excel_template_path)
    
    sheets_to_process = []
    if export_type == "TOUS":
        sheets_to_process = [("CARBURANT", 7, 4, 3), ("EAU", 11, 5, 4)]
    elif export_type == "CARBURANT":
        sheets_to_process = [("CARBURANT", 7, 4, 3)]
        if "EAU" in wb.sheetnames:
            wb.remove(wb["EAU"])
    elif export_type == "EAU":
        sheets_to_process = [("EAU", 11, 5, 4)]
        if "CARBURANT" in wb.sheetnames:
            wb.remove(wb["CARBURANT"])
            
    # Styling definitions
    thin_border = Border(
        left=Side(style='thin', color='C0C0C0'),
        right=Side(style='thin', color='C0C0C0'),
        top=Side(style='thin', color='C0C0C0'),
        bottom=Side(style='thin', color='C0C0C0')
    )
    
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(name="Calibri", size=9, bold=True, color="006100")
    
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    yellow_font = Font(name="Calibri", size=9, bold=True, color="9C6500")
    
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    white_font = Font(name="Calibri", size=9, color="808080")
    
    for sheet_name, start_row, code_col, comment_col in sheets_to_process:
        ws = wb[sheet_name]
        
        citernes = get_all_citernes(sheet_name)
        etapes = get_etapes(sheet_name)
        
        # Read current status from DB
        conn = get_connection()
        query = """
        SELECT citerne_code, step_id, completion_pct
        FROM suivi_progress
        """
        progress_df = pd.read_sql_query(query, conn)
        conn.close()
        
        # Map step_id to step_order
        etape_id_to_order = dict(zip(etapes['id'], etapes['step_order']))
        
        # Loop through rows starting from start_row
        row_idx = start_row
        for i, cit_row in citernes.iterrows():
            citerne_code = cit_row['code']
            citerne_comment = cit_row['comments']
            
            # Write Citerne Code
            write_cell_safely(ws, row_idx, code_col, citerne_code, font=Font(name="Calibri", size=9, bold=True))
            
            # Write Comment
            write_cell_safely(ws, row_idx, comment_col, citerne_comment, font=Font(name="Calibri", size=9))
            
            # Get progress for this Citerne
            cit_prog = progress_df[progress_df['citerne_code'] == citerne_code]
            prog_dict = {}
            for _, cp_row in cit_prog.iterrows():
                step_order = etape_id_to_order.get(cp_row['step_id'])
                if step_order:
                    prog_dict[step_order] = cp_row['completion_pct']
            
            # Write values for each step column
            for step_order in range(1, len(etapes) + 1):
                col_idx = code_col + step_order
                pct = prog_dict.get(step_order, 0.0)
                
                # Format cell value (1 for 100%, 0.8 for 80%, etc.)
                if pct >= 99.9:
                    val = 1.0
                    fill = green_fill
                    font = green_font
                elif pct > 0:
                    val = float(pct / 100.0)
                    fill = yellow_fill
                    font = yellow_font
                else:
                    val = 0.0
                    fill = white_fill
                    font = white_font
                    
                write_cell_safely(
                    ws, 
                    row_idx, 
                    col_idx, 
                    val, 
                    fill=fill, 
                    font=font, 
                    alignment=Alignment(horizontal='center', vertical='center'),
                    border=thin_border,
                    number_format='0%'
                )
                
            row_idx += 1
            
    output = python_io.BytesIO()
    wb.save(output)
    return output.getvalue()
